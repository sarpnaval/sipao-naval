"""SARP-Naval · Importador de datos SISLOG (tarea 2.3).

Lee la plantilla de importación de 03-datos/plantillas/ (formato de
README_DATOS.md), valida TODOS los errores (no solo el primero), y —solo
si el usuario lo pide y no hay errores— aplica la importación en UNA
transacción SQLite que reemplaza los datos y recalcula parámetros,
pronósticos, alertas y clasificación con backend.motor.analisis_real.

Principios (CLAUDE.md):
- La app NUNCA escribe en SISLOG: aquí solo se LEEN archivos que el
  usuario sube (exportados de SISLOG por el canal formal).
- SQL siempre parametrizado. Sin dependencias de red en tiempo de
  ejecución (openpyxl y csv trabajan sobre bytes en memoria).
- Aplicación ATÓMICA: si algo falla a mitad, rollback y la base
  anterior queda intacta.

Formatos aceptados:
- Un .xlsx con tres hojas llamadas maestro_items, movimientos y
  stock_actual (se leen con openpyxl en modo solo lectura).
- Tres .csv (uno por tabla); cada uno se reconoce por sus encabezados,
  sin importar el nombre del archivo. Se tolera BOM UTF-8, codificación
  utf-8 o latin-1 (se prueba utf-8 primero; latin-1 nunca falla y actúa
  de red de seguridad), separador coma o punto y coma (se detecta
  contando cuál aparece más en la línea de encabezados) y espacios
  alrededor de cada celda. Las líneas cuya primera celda empieza con
  '#' son comentarios y se ignoran (así las plantillas publicadas
  llevan filas de ejemplo documentadas).

Fechas aceptadas: dd/mm/aaaa o yyyy-mm-dd (y celdas fecha nativas de
Excel). Números: punto decimal, y también coma decimal simple ("3,5").
"""

import csv
import datetime
import io
import re
from collections import Counter

from backend.app.fechas import fecha_larga
from backend.motor.analisis_real import (
    MESES_MINIMO_UTIL,
    MESES_POLITICA_MINIMOS,
    analizar_importacion,
)
from backend.motor.motor import Z as Z_SERVICIO

__all__ = ["importar", "leer_archivos", "validar", "aplicar_importacion",
           "descripcion_plantilla", "VERSION_MODELO"]

VERSION_MODELO = "v1.0-import"

# Encabezados por tabla: (obligatorios, opcionales). Nombres tal cual
# los define la plantilla de README_DATOS.md (minúsculas, sin espacios).
ENCABEZADOS = {
    "maestro_items": (
        ["codigo", "nombre", "categoria", "unidad", "costo_unitario_usd",
         "criticidad", "lead_time_dias", "importado"],
        ["proveedor"],
    ),
    "movimientos": (
        ["codigo", "fecha", "tipo", "cantidad", "reparto"],
        ["referencia"],
    ),
    "stock_actual": (
        ["codigo", "existencia", "fecha_corte"],
        ["ubicacion"],
    ),
}

TIPOS_MOVIMIENTO = {"CONSUMO", "INGRESO", "AJUSTE"}

_RE_FECHA_DMA = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")
_RE_FECHA_ISO = re.compile(r"^(\d{4})-(\d{1,2})-(\d{1,2})$")


# ---------------------------------------------------------------------
# Utilidades de parseo de celdas
# ---------------------------------------------------------------------
def _texto(valor):
    """Celda como texto limpio ('' si viene vacía o None)."""
    if valor is None:
        return ""
    return str(valor).strip()


def _numero(valor):
    """Convierte una celda a float; None si no es numérica.

    Acepta int/float nativos (celdas de Excel) y texto con punto
    decimal o coma decimal simple ('3,5' → 3.5).
    """
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = _texto(valor)
    if not texto:
        return None
    try:
        return float(texto)
    except ValueError:
        if texto.count(",") == 1 and "." not in texto:
            try:
                return float(texto.replace(",", "."))
            except ValueError:
                return None
        return None


def _fecha_iso(valor):
    """Convierte una celda de fecha a ISO yyyy-mm-dd; None si inválida.

    Acepta datetime/date nativos (Excel), 'dd/mm/aaaa' y 'yyyy-mm-dd'.
    Verifica que la fecha exista en el calendario (31/02 es inválido).
    """
    if isinstance(valor, datetime.datetime):
        return valor.date().isoformat()
    if isinstance(valor, datetime.date):
        return valor.isoformat()
    texto = _texto(valor)
    m = _RE_FECHA_DMA.match(texto)
    if m:
        dia, mes, anio = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = _RE_FECHA_ISO.match(texto)
        if not m:
            return None
        anio, mes, dia = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime.date(anio, mes, dia).isoformat()
    except ValueError:
        return None


def _normalizar_encabezado(valor):
    """Encabezado en minúsculas, sin BOM ni espacios."""
    return _texto(valor).lstrip("﻿").strip().lower()


# ---------------------------------------------------------------------
# Lectura de archivos (CSV y XLSX) → tablas crudas
# ---------------------------------------------------------------------
def _decodificar_csv(contenido):
    """Decodifica bytes CSV: utf-8 (con o sin BOM) y si no, latin-1."""
    try:
        return contenido.decode("utf-8-sig")
    except UnicodeDecodeError:
        return contenido.decode("latin-1")


def _detectar_separador(texto):
    """Detecta ',' o ';' contando cuál domina en la línea de encabezados."""
    primera_linea = texto.split("\n", 1)[0]
    return ";" if primera_linea.count(";") > primera_linea.count(",") else ","


def _filas_desde_matriz(matriz):
    """Convierte una matriz (encabezado + filas) en filas normalizadas.

    Devuelve (encabezados, [(numero_fila, {columna: valor})]). Se saltan
    filas vacías y comentarios (primera celda que empieza con '#').
    """
    encabezados = []
    filas = []
    for numero, celdas in enumerate(matriz, start=1):
        if not encabezados:
            encabezados = [_normalizar_encabezado(c) for c in celdas
                           if _normalizar_encabezado(c)]
            continue
        primera = _texto(celdas[0]) if celdas else ""
        if primera.startswith("#"):
            continue
        if all(_texto(c) == "" for c in celdas):
            continue
        fila = {}
        for i, columna in enumerate(encabezados):
            fila[columna] = celdas[i] if i < len(celdas) else None
        filas.append((numero, fila))
    return encabezados, filas


def _clasificar_tabla(encabezados):
    """Identifica a qué tabla corresponde un CSV según sus encabezados."""
    conjunto = set(encabezados)
    if "costo_unitario_usd" in conjunto:
        return "maestro_items"
    if "tipo" in conjunto and "fecha" in conjunto:
        return "movimientos"
    if "existencia" in conjunto:
        return "stock_actual"
    return None


def leer_archivos(archivos):
    """Lee la lista [(nombre, bytes)] y devuelve (tablas, errores).

    tablas: {nombre_tabla: {"origen": str, "encabezados": [...],
    "filas": [(numero_fila, dict)]}}. Los errores de lectura (extensión
    no soportada, hoja faltante, tabla no reconocida o duplicada) se
    reportan todos, con el mismo formato que los de validación.
    """
    tablas = {}
    errores = []

    def _registrar(nombre_tabla, origen, encabezados, filas):
        if nombre_tabla in tablas:
            errores.append(_error(
                origen, None, None,
                f"La tabla '{nombre_tabla}' ya fue aportada por "
                f"'{tablas[nombre_tabla]['origen']}': archivo duplicado."))
            return
        tablas[nombre_tabla] = {"origen": origen,
                                "encabezados": encabezados, "filas": filas}

    for nombre, contenido in archivos:
        nombre_bajo = (nombre or "archivo").lower()
        if nombre_bajo.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
                libro = load_workbook(io.BytesIO(contenido),
                                      read_only=True, data_only=True)
            except Exception:
                errores.append(_error(
                    nombre, None, None,
                    "No se pudo abrir el archivo .xlsx: verifique que sea "
                    "un libro de Excel válido."))
                continue
            for hoja in ENCABEZADOS:
                if hoja not in libro.sheetnames:
                    errores.append(_error(
                        nombre, None, None,
                        f"Falta la hoja obligatoria '{hoja}' en el libro "
                        f"(hojas encontradas: {', '.join(libro.sheetnames)})."))
                    continue
                matriz = [list(f) for f in
                          libro[hoja].iter_rows(values_only=True)]
                encabezados, filas = _filas_desde_matriz(matriz)
                _registrar(hoja, f"{nombre} · hoja {hoja}", encabezados, filas)
            libro.close()
        elif nombre_bajo.endswith(".csv"):
            texto = _decodificar_csv(contenido)
            separador = _detectar_separador(texto)
            matriz = list(csv.reader(io.StringIO(texto),
                                     delimiter=separador))
            encabezados, filas = _filas_desde_matriz(matriz)
            tabla = _clasificar_tabla(encabezados)
            if tabla is None:
                errores.append(_error(
                    nombre, None, None,
                    "No se reconoce la tabla por sus encabezados. Se "
                    "esperan los de maestro_items, movimientos o "
                    "stock_actual (ver GET /api/importar/plantilla)."))
                continue
            _registrar(tabla, nombre, encabezados, filas)
        else:
            errores.append(_error(
                nombre, None, None,
                "Extensión no soportada: solo se aceptan .xlsx (3 hojas) "
                "o .csv (uno por tabla)."))

    for tabla in ENCABEZADOS:
        if tabla not in tablas:
            errores.append(_error(
                None, None, None,
                f"No se encontró la tabla '{tabla}' entre los archivos "
                "subidos: la importación necesita las tres tablas."))
    return tablas, errores


# ---------------------------------------------------------------------
# Validación (reporta TODOS los errores)
# ---------------------------------------------------------------------
def _error(archivo, fila, columna, mensaje):
    """Estructura homogénea de error para el reporte en pantalla."""
    return {"archivo": archivo, "fila": fila, "columna": columna,
            "mensaje": mensaje}


def _validar_encabezados(tabla, datos_tabla, errores):
    """Verifica encabezados obligatorios; True si la tabla es usable."""
    obligatorios, _ = ENCABEZADOS[tabla]
    faltantes = [c for c in obligatorios
                 if c not in datos_tabla["encabezados"]]
    for columna in faltantes:
        errores.append(_error(
            datos_tabla["origen"], 1, columna,
            f"Falta la columna obligatoria '{columna}' en {tabla}."))
    return not faltantes


def validar(tablas):
    """Valida las tres tablas y devuelve (datos, errores, advertencias).

    Reporta TODOS los errores encontrados (no se detiene en el primero),
    cada uno con archivo/hoja, fila, columna y mensaje en español.
    `datos` (solo confiable si no hay errores) contiene maestro,
    movimientos, stock, reparto detectado y fecha de corte.
    """
    errores = []
    advertencias = []
    maestro = []
    movimientos = []
    stock = {}

    # ---- maestro_items ----
    codigos_maestro = set()
    tabla = tablas.get("maestro_items")
    if tabla and _validar_encabezados("maestro_items", tabla, errores):
        origen = tabla["origen"]
        if not tabla["filas"]:
            errores.append(_error(origen, None, None,
                                  "El maestro de ítems no contiene filas."))
        for numero, fila in tabla["filas"]:
            codigo = _texto(fila.get("codigo"))
            if not codigo:
                errores.append(_error(origen, numero, "codigo",
                                      "El código del ítem está vacío."))
            elif codigo in codigos_maestro:
                errores.append(_error(
                    origen, numero, "codigo",
                    f"Código duplicado en el maestro: '{codigo}'."))
            else:
                codigos_maestro.add(codigo)

            for columna in ("nombre", "categoria", "unidad"):
                if not _texto(fila.get(columna)):
                    errores.append(_error(
                        origen, numero, columna,
                        f"La columna '{columna}' está vacía."))

            costo = _numero(fila.get("costo_unitario_usd"))
            if costo is None or costo <= 0:
                errores.append(_error(
                    origen, numero, "costo_unitario_usd",
                    "El costo unitario debe ser un número mayor que 0 "
                    f"(valor recibido: '{_texto(fila.get('costo_unitario_usd'))}')."))

            criticidad = _texto(fila.get("criticidad")).upper()
            if criticidad not in ("V", "E", "D"):
                errores.append(_error(
                    origen, numero, "criticidad",
                    "Criticidad inválida: debe ser V (vital), E (esencial) "
                    f"o D (deseable); se recibió '{_texto(fila.get('criticidad'))}'."))

            lead_time = _numero(fila.get("lead_time_dias"))
            if (lead_time is None or lead_time <= 0
                    or not float(lead_time).is_integer()):
                errores.append(_error(
                    origen, numero, "lead_time_dias",
                    "El lead time debe ser un entero de días mayor que 0 "
                    f"(valor recibido: '{_texto(fila.get('lead_time_dias'))}')."))

            importado = _texto(fila.get("importado")).upper()
            if importado not in ("S", "N"):
                errores.append(_error(
                    origen, numero, "importado",
                    "La columna importado debe ser S o N; se recibió "
                    f"'{_texto(fila.get('importado'))}'."))

            maestro.append({
                "codigo": codigo,
                "nombre": _texto(fila.get("nombre")),
                "categoria": _texto(fila.get("categoria")),
                "unidad": _texto(fila.get("unidad")),
                "costo": costo,
                "crit": criticidad if criticidad in ("V", "E", "D") else "D",
                "lt": int(lead_time) if lead_time and lead_time > 0
                      and float(lead_time).is_integer() else 1,
                "imp": importado == "S",
                "proveedor": _texto(fila.get("proveedor")) or None,
            })

    # ---- movimientos ----
    tabla = tablas.get("movimientos")
    vistos = set()  # duplicados exactos
    if tabla and _validar_encabezados("movimientos", tabla, errores):
        origen = tabla["origen"]
        if not tabla["filas"]:
            errores.append(_error(
                origen, None, None,
                "La tabla de movimientos no contiene filas: sin historia "
                "no hay nada que analizar."))
        for numero, fila in tabla["filas"]:
            codigo = _texto(fila.get("codigo"))
            if not codigo:
                errores.append(_error(origen, numero, "codigo",
                                      "El código del movimiento está vacío."))
            elif codigos_maestro and codigo not in codigos_maestro:
                errores.append(_error(
                    origen, numero, "codigo",
                    f"El código '{codigo}' no existe en el maestro de ítems."))

            fecha = _fecha_iso(fila.get("fecha"))
            if fecha is None:
                errores.append(_error(
                    origen, numero, "fecha",
                    "Fecha inválida: use dd/mm/aaaa o yyyy-mm-dd "
                    f"(valor recibido: '{_texto(fila.get('fecha'))}')."))

            tipo = _texto(fila.get("tipo")).upper()
            if tipo not in TIPOS_MOVIMIENTO:
                errores.append(_error(
                    origen, numero, "tipo",
                    "Tipo de movimiento inválido: debe ser CONSUMO, "
                    f"INGRESO o AJUSTE; se recibió '{_texto(fila.get('tipo'))}'."))

            cantidad = _numero(fila.get("cantidad"))
            if cantidad is None:
                errores.append(_error(
                    origen, numero, "cantidad",
                    "La cantidad debe ser numérica (valor recibido: "
                    f"'{_texto(fila.get('cantidad'))}')."))
            elif tipo == "CONSUMO" and cantidad < 0:
                errores.append(_error(
                    origen, numero, "cantidad",
                    f"Cantidad negativa en un CONSUMO ({cantidad:g}): "
                    "los consumos deben ser mayores o iguales a 0."))

            reparto = _texto(fila.get("reparto"))
            if not reparto:
                errores.append(_error(
                    origen, numero, "reparto",
                    "La columna reparto está vacía."))

            clave = (codigo, fecha, tipo,
                     cantidad, reparto, _texto(fila.get("referencia")))
            if fecha is not None and cantidad is not None:
                if clave in vistos:
                    errores.append(_error(
                        origen, numero, None,
                        f"Movimiento duplicado exacto: {codigo} {tipo} "
                        f"{cantidad:g} del {fecha_larga(fecha)} ya aparece "
                        "en una fila anterior."))
                vistos.add(clave)

            movimientos.append({
                "codigo": codigo,
                "fecha": fecha,
                "tipo": tipo,
                "cantidad": cantidad,
                "reparto": reparto,
                "referencia": _texto(fila.get("referencia")) or None,
            })

    # ---- stock_actual ----
    tabla = tablas.get("stock_actual")
    if tabla and _validar_encabezados("stock_actual", tabla, errores):
        origen = tabla["origen"]
        for numero, fila in tabla["filas"]:
            codigo = _texto(fila.get("codigo"))
            if not codigo:
                errores.append(_error(origen, numero, "codigo",
                                      "El código de la fila de stock está vacío."))
            elif codigos_maestro and codigo not in codigos_maestro:
                errores.append(_error(
                    origen, numero, "codigo",
                    f"El código '{codigo}' no existe en el maestro de ítems."))
            elif codigo in stock:
                errores.append(_error(
                    origen, numero, "codigo",
                    f"Código repetido en stock_actual: '{codigo}' (solo se "
                    "admite una fila de existencia por ítem)."))

            existencia = _numero(fila.get("existencia"))
            if existencia is None or existencia < 0:
                errores.append(_error(
                    origen, numero, "existencia",
                    "La existencia debe ser un número mayor o igual a 0 "
                    f"(valor recibido: '{_texto(fila.get('existencia'))}')."))

            fecha_corte = _fecha_iso(fila.get("fecha_corte"))
            if fecha_corte is None:
                errores.append(_error(
                    origen, numero, "fecha_corte",
                    "Fecha de corte inválida: use dd/mm/aaaa o yyyy-mm-dd "
                    f"(valor recibido: '{_texto(fila.get('fecha_corte'))}')."))

            if codigo and codigo not in stock:
                stock[codigo] = {
                    "existencia": existencia if existencia is not None else 0.0,
                    "fecha_corte": fecha_corte,
                    "ubicacion": _texto(fila.get("ubicacion")) or None,
                }

    # ---- advertencias estructurales (no bloquean) ----
    codigos_con_consumo = {m["codigo"] for m in movimientos
                           if m["tipo"] == "CONSUMO"}
    for it in maestro:
        if it["codigo"] and it["codigo"] not in codigos_con_consumo:
            advertencias.append(
                f"El ítem '{it['codigo']}' no tiene movimientos de consumo: "
                "entrará en política de mínimos sin base de demanda.")
        if it["codigo"] and it["codigo"] not in stock:
            advertencias.append(
                f"El ítem '{it['codigo']}' no tiene fila en stock_actual: "
                "se asume existencia 0.")

    reparto = None
    repartos = [m["reparto"] for m in movimientos if m["reparto"]]
    if repartos:
        reparto = Counter(repartos).most_common(1)[0][0]
    cortes = [s["fecha_corte"] for s in stock.values() if s["fecha_corte"]]
    fechas_mov = [m["fecha"] for m in movimientos if m["fecha"]]
    fecha_corte = max(cortes) if cortes else (max(fechas_mov)
                                              if fechas_mov else None)

    datos = {"maestro": maestro, "movimientos": movimientos, "stock": stock,
             "reparto": reparto, "fecha_corte": fecha_corte}
    return datos, errores, advertencias


# ---------------------------------------------------------------------
# Aplicación atómica a la base
# ---------------------------------------------------------------------
# Orden de borrado que respeta las claves foráneas (hijas primero)
_TABLAS_A_REEMPLAZAR = ("alertas", "pronosticos", "parametros",
                        "clasificacion", "stock", "movimientos",
                        "items", "metadatos")


def aplicar_importacion(conexion, datos, analisis, fecha_importacion=None,
                        origen="importado"):
    """Reemplaza los datos y resultados en UNA transacción SQLite.

    Si cualquier sentencia falla, se hace rollback y la base anterior
    queda intacta (sqlite3 abre la transacción implícita en el primer
    DML y solo la cierra el commit final). `origen` permite al registro
    directo preservar el origen previo del dataset (p. ej.
    'simulado-realista') dentro del MISMO COMMIT, sin transacciones
    posteriores.
    """
    from backend.app.seed import _priorizar_alertas  # reutiliza dossier §5.4

    if fecha_importacion is None:
        fecha_importacion = datetime.date.today().isoformat()
    fecha_corte = analisis["fecha_corte"]
    items = analisis["items"]

    cursor = conexion.cursor()
    try:
        for tabla in _TABLAS_A_REEMPLAZAR:
            cursor.execute(f"DELETE FROM {tabla}")  # nombres fijos, no de usuario

        cursor.executemany(
            """INSERT INTO items (codigo, nombre, categoria, unidad,
                   costo_unitario, criticidad_ved, lead_time_dias,
                   importado, proveedor)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [(f["codigo"], f["nombre"], f["categoria"], f["unidad"],
              f["costo"], f["crit"], f["lt"], 1 if f["imp"] else 0,
              f["proveedor"]) for f in datos["maestro"]],
        )

        # Movimientos CRUDOS tal como vinieron (tipo en minúsculas por el
        # CHECK del esquema; fecha ya normalizada a ISO)
        cursor.executemany(
            """INSERT INTO movimientos (codigo_item, fecha, tipo, cantidad,
                   reparto, orden_ref)
               VALUES (?, ?, ?, ?, ?, ?)""",
            [(m["codigo"], m["fecha"], m["tipo"].lower(), m["cantidad"],
              m["reparto"], m["referencia"]) for m in datos["movimientos"]],
        )

        # Stock: una fila por ítem del maestro; sin fila → existencia 0
        filas_stock = []
        for f in datos["maestro"]:
            fila = datos["stock"].get(f["codigo"])
            filas_stock.append((
                f["codigo"], analisis["reparto"],
                fila["existencia"] if fila else 0.0,
                (fila["fecha_corte"] if fila and fila["fecha_corte"]
                 else fecha_corte),
                fila["ubicacion"] if fila else None,
            ))
        cursor.executemany(
            """INSERT INTO stock (codigo_item, reparto, existencia,
                   fecha_corte, ubicacion)
               VALUES (?, ?, ?, ?, ?)""",
            filas_stock,
        )

        # Parámetros: la política de mínimos queda auditada en
        # version_modelo (sin tocar el esquema)
        cursor.executemany(
            """INSERT INTO parametros (codigo_item, z_servicio, ss, rop,
                   eoq, nivel_max, fecha_calculo, version_modelo)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            [(it["id"], Z_SERVICIO[it["crit"]], it["ss"], it["rop"],
              it["eoq"], it["maxLevel"], fecha_importacion,
              VERSION_MODELO + ("-minimos" if it["politica_minimos"] else ""))
             for it in items],
        )

        # Pronósticos: SOLO los estadísticos (holt/croston). Los ítems en
        # política de mínimos no tienen pronóstico estadístico (honesto
        # con el dossier §5.5 y con el CHECK del esquema).
        filas_pron = []
        for it in items:
            if it["politica_minimos"]:
                continue
            for k, prevision in enumerate(it["forecast"]):
                filas_pron.append((it["id"], analisis["fcLabels"][k],
                                   prevision, it["sigma"], it["mape"],
                                   it["modelo"]))
        cursor.executemany(
            """INSERT INTO pronosticos (codigo_item, mes, demanda_prevista,
                   sigma, mape, modelo)
               VALUES (?, ?, ?, ?, ?, ?)""",
            filas_pron,
        )

        # Alertas priorizadas según dossier §5.4 (misma regla que seed)
        for prioridad, it in enumerate(_priorizar_alertas(items), start=1):
            cursor.execute(
                """INSERT INTO alertas (codigo_item, estado, dias_a_quiebre,
                       cantidad_sugerida, prioridad, fecha, atendida)
                   VALUES (?, ?, ?, ?, ?, ?, 0)""",
                (it["id"], it["estado"], it["diasQuiebre"], it["sugerido"],
                 prioridad, fecha_corte),
            )

        cursor.executemany(
            """INSERT INTO clasificacion (codigo_item, abc, xyz, cv,
                   valor_anual, valor_stock, dias_quiebre, demanda_mensual)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            [(it["id"], it["abc"], it["xyz"], it["cv"], it["valorAnual"],
              it["valorStock"], it["diasQuiebre"], it["dAvg"])
             for it in items],
        )

        cursor.executemany(
            "INSERT INTO metadatos (clave, valor) VALUES (?, ?)",
            [("generado", fecha_larga(fecha_corte)),
             ("reparto", analisis["reparto"]),
             ("fecha_corte", fecha_corte),
             ("version_modelo", VERSION_MODELO),
             ("origen", origen),
             ("fecha_importacion", fecha_importacion),
             ("fecha_datos", fecha_corte)],
        )

        conexion.commit()
    except Exception:
        conexion.rollback()
        raise


# ---------------------------------------------------------------------
# Orquestador: validar (dry-run) y opcionalmente aplicar
# ---------------------------------------------------------------------
def importar(archivos, conexion=None, aplicar=False, fecha_importacion=None):
    """Importa (o solo valida) los archivos subidos.

    - archivos: lista de tuplas (nombre, bytes).
    - conexion: conexión SQLite abierta (obligatoria si aplicar=True).
    - aplicar=False (por defecto): DRY-RUN, valida y analiza sin tocar
      NUNCA la base de datos.
    - aplicar=True: si no hay errores, reemplaza los datos en una
      transacción atómica y recalcula todo.

    Devuelve el reporte: {valido, errores, advertencias, resumen,
    aplicado, kpi, calidad_datos}.
    """
    tablas, errores_lectura = leer_archivos(archivos)
    datos, errores_validacion, advertencias = validar(tablas)
    errores = errores_lectura + errores_validacion

    reporte = {
        "valido": not errores,
        "errores": errores,
        "advertencias": advertencias,
        "resumen": {
            "items": len(datos["maestro"]),
            "movimientos": len(datos["movimientos"]),
            "meses_historia": None,
            "rango_fechas": None,
            "reparto": datos["reparto"],
        },
        "aplicado": False,
        "kpi": None,
        "calidad_datos": None,
    }
    if errores:
        return reporte

    analisis = analizar_importacion(
        datos["maestro"], datos["movimientos"], datos["stock"],
        reparto=datos["reparto"], fecha_corte=datos["fecha_corte"])

    reporte["resumen"]["meses_historia"] = analisis["meses_historia"]
    reporte["resumen"]["rango_fechas"] = {
        "desde": fecha_larga(analisis["rango_fechas"]["desde"]),
        "hasta": fecha_larga(analisis["rango_fechas"]["hasta"]),
    }
    reporte["resumen"]["reparto"] = analisis["reparto"]
    reporte["calidad_datos"] = analisis["calidad_datos"]

    # Advertencias de historia (dossier §5.5 y README_DATOS regla 2)
    for registro in analisis["calidad_datos"]["historia_corta"]:
        advertencias.append(
            f"El ítem '{registro['codigo']}' tiene solo {registro['meses']} "
            f"meses de historia (< {MESES_POLITICA_MINIMOS}): opera en "
            "política de mínimos hasta acumular datos (dossier §5.5).")
    for registro in analisis["calidad_datos"]["historia_bajo_minimo"]:
        advertencias.append(
            f"El ítem '{registro['codigo']}' tiene {registro['meses']} "
            f"meses de historia, bajo el mínimo útil de {MESES_MINIMO_UTIL} "
            "meses recomendado (README_DATOS regla 2).")

    if aplicar:
        if conexion is None:
            raise ValueError(
                "Se pidió aplicar la importación sin una conexión a la base.")
        aplicar_importacion(conexion, datos, analisis, fecha_importacion)
        reporte["aplicado"] = True
        reporte["kpi"] = analisis["kpi"]

    return reporte


# ---------------------------------------------------------------------
# Descripción de la plantilla (para GET /api/importar/plantilla)
# ---------------------------------------------------------------------
def descripcion_plantilla():
    """Descripción del formato de importación para mostrar en pantalla."""
    return {
        "descripcion": (
            "Plantilla de importación SARP-Naval (README_DATOS.md). Se "
            "acepta UN archivo .xlsx con tres hojas (maestro_items, "
            "movimientos, stock_actual) o TRES archivos .csv, uno por "
            "tabla, reconocidos por sus encabezados."),
        "formatos": {
            "fechas": ["dd/mm/aaaa", "yyyy-mm-dd"],
            "separador_csv": [",", ";"],
            "codificacion_csv": ["UTF-8 (con o sin BOM)", "Latin-1"],
            "comentarios": "Las filas cuya primera celda empieza con '#' se ignoran.",
        },
        "tablas": {
            "maestro_items": {
                "columnas_obligatorias": ENCABEZADOS["maestro_items"][0],
                "columnas_opcionales": ENCABEZADOS["maestro_items"][1],
                "ejemplo": {
                    "codigo": "REP-0001",
                    "nombre": "Filtro de combustible motor F/B 200HP",
                    "categoria": "Repuestos motor",
                    "unidad": "unidad",
                    "costo_unitario_usd": 38.50,
                    "criticidad": "V",
                    "lead_time_dias": 45,
                    "importado": "S",
                    "proveedor": "Proveedor Ejemplo S.A.",
                },
            },
            "movimientos": {
                "columnas_obligatorias": ENCABEZADOS["movimientos"][0],
                "columnas_opcionales": ENCABEZADOS["movimientos"][1],
                "historia_recomendada_meses": "24 a 36",
                "ejemplo": {
                    "codigo": "REP-0001",
                    "fecha": "15/01/2025",
                    "tipo": "CONSUMO",
                    "cantidad": 6,
                    "reparto": "Reparto guardacostas SUBNOR (referencial)",
                    "referencia": "OT-2025-014",
                },
            },
            "stock_actual": {
                "columnas_obligatorias": ENCABEZADOS["stock_actual"][0],
                "columnas_opcionales": ENCABEZADOS["stock_actual"][1],
                "ejemplo": {
                    "codigo": "REP-0001",
                    "existencia": 12,
                    "fecha_corte": "30/06/2026",
                    "ubicacion": "Bodega A · Estante 3",
                },
            },
        },
        "reglas": [
            "Criticidad: V (vital), E (esencial) o D (deseable).",
            "Importado: S o N (define el costo de emisión de orden: "
            "USD 120 importado, USD 45 nacional).",
            "Cantidades de CONSUMO mayores o iguales a 0.",
            "Todo código de movimientos y stock debe existir en el maestro.",
            "Ítems con historia menor a 18 meses operan en política de "
            "mínimos (dossier §5.5); el mínimo útil son 24 meses.",
        ],
    }
